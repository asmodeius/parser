import openpyxl
from pathlib import Path
#open excel file
wb_obj = openpyxl.load_workbook("data.xlsx") 

# Read the active sheet:
sheet = wb_obj.active
result = []
index = 0
while(True):
	index = index + 1
	# if we got to empty row we are done
	if sheet.cell(row=index, column=1).value is None:
		break
	# check if transfer is incoming and it is purchase
	if sheet.cell(row=index, column=3).value == "incoming" and sheet.cell(row=index, column=4).value == "purchase":
		# get payed amount
		cash = int(sheet.cell(row=index, column=5).value)
		# price of ticket
		price = 500
		# this can be more advanced but I would need data for that
		comment = sheet.cell(row=index, column=6).value[27:]
		# calculate number of payed tickets from payed amount and price of ticket
		tickets = int(cash/price)
		# add to list tickets bought
		for i in range(tickets):
			result.append(comment)

# open file to save the result into
f = open("result.txt", "a")
for el in result:
	# we print
	print(el)
	# and save to file
	f.write(el+'\n')

